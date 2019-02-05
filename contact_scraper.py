#! python3.7
# author: Stanislav Matas
# date: 03/02/2019
# name: Selenium web scraper originally for MH marketing to collect contact data into a database about estates
# to be rent in selected areas of Spain, France, Germany, Italy, England and so on.

# How to use
# 0. the input URL MUST be always the first URL of the listing (page number 1)
# 1. when neither from_page nor to_page is input, run it from the first page to the last page that was found
# 2. when only from_page was input (!with keyword argument!), run it from from_page to the last page that was found
# 3. when only to_page was input, run it from the first page to the to_page but:
#      a) to_page cannot exceed last existing page, it is lowered to the last page that was found if so
#      b) if to_page=1 was input, only 1 page is scraped and the program terminates if so
# 4. when from_page and to_page were input, run from from_page to to_page but:
#      a) to_page cannot exceed last existing page, it is lowered to the last page that was found if so
#      b) to_page must be a bigger number than from_page, a warning appears and the program terminates if so
#      c) if from_page and to_page are the same numbers, only 1 page is scraped and the program terminates if so


import multiprocessing as mp
import datetime
import time
import json
import os
import sys

import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.firefox.firefox_profile import AddonFormatError, FirefoxProfile
from selenium.webdriver.firefox.options import Options
from selenium.common.exceptions import NoSuchElementException, TimeoutException, InvalidElementStateException

import sl_selectors

conn, cur = None, None
db_connected = False


def connect_db():
    """Create a connection to a database located at localhost."""

    global conn, cur, db_connected
    conn = pymysql.connect(host='localhost', unix_socket='/run/mysqld/mysqld.sock',
                           user='mh_selenium', passwd='mh_selenium', db='mh', charset='utf8')
    cur = conn.cursor()
    cur.execute('USE mh')
    db_connected = True
    print('--- Database connection established ---\n')


def disconnect_db():
    """Safely close a connection to a database if existed."""

    global conn, cur, db_connected
    if conn and cur:
        cur.close()
        conn.close()
        conn, cur = None, None
        db_connected = False
        print('--- Database connection closed ---\n')


class NoAdsFound(Exception):
    pass


# in case of installing modern extensions:
class FirefoxProfileWithWebExtensionSupport(webdriver.FirefoxProfile):
    def _addon_details(self, addon_path):
        try:
            return super()._addon_details(addon_path)
        except AddonFormatError:
            try:
                with open(os.path.join(addon_path, 'manifest.json'), 'r') as f:
                    manifest = json.load(f)
                    return {
                        'id': manifest['applications']['gecko']['id'],
                        'version': manifest['version'],
                        'name': manifest['name'],
                        'unpack': False,
                    }
            except (IOError, KeyError) as e:
                raise AddonFormatError(str(e), sys.exc_info()[2])


def open_firefox_exts_headless() -> webdriver:
    """Launch Firefox in headless mode and enable add-ons.

    :return: Instance of Firefox Selenium webdriver.
    """

    print('Launching Firefox in headless mode...')
    profile = FirefoxProfile(profile_directory='/home/standa/.mozilla/firefox/mrgxem71.scraping/')
    options = Options()
    options.headless = True
    driver = webdriver.Firefox(firefox_profile=profile, options=options)
    time.sleep(2)
    while True:
        try:
            driver.get('about:addons')
            driver.find_element_by_name('NoScript').click()
            driver.find_element_by_id('detail-enable-btn').click()
            print(' - NoScript add-on enabled.')
            driver.back()
            driver.find_element_by_name('uBlock Origin').click()
            driver.find_element_by_id('detail-enable-btn').click()
            print(' - uBlock Origin add-on enabled.\n')
            break
        except (NoSuchElementException, InvalidElementStateException):
            print('Could not enable the extensions. Retrying...')
            time.sleep(2)
    driver.set_page_load_timeout(30)
    return driver


def recognize_sln_selectors(driver: webdriver = None, first_listing_url: str = None,
                            from_page: int = None, to_page: int = None) -> sl_selectors:
    """Based on the input first listing URL, return the corresponding set of Selenium selectors for data scraping."""

    websites = {'https://www.traum-ferienwohnungen.de':
                sl_selectors.TraumFerienWohnungen(driver, first_listing_url, from_page, to_page),
                'https://www.milanuncios.com':
                sl_selectors.MilAnuncios(driver, first_listing_url, from_page, to_page),
                'https://www.vivaweek.com':
                sl_selectors.VivaWeek(driver, first_listing_url, from_page, to_page),
                'https://vacances.seloger.com':
                sl_selectors.VacancesSeloger(driver, first_listing_url, from_page, to_page)
                }

    for website, sln_selector in websites.items():
        if driver and driver.current_url.startswith(website):
            return sln_selector
        elif first_listing_url and first_listing_url.startswith(website):
            return sln_selector


def evaluate_paging(from_page: int = None, to_page: int = None) -> None:
    """Print out a warning in case of starting the Selenium scraper with incorrect arguments and terminate."""

    if to_page and from_page:
        if to_page < from_page:
            print(f'Set the upper limit ({to_page}) higher than the lower limit ({from_page}) '
                  f'of the listing pages to scrape.\n'
                  'Terminating...')
            sys.exit()
    if from_page and not to_page:
        print('Set also the upper limit of the listing pages to scrape.\n'
              'Terminating...')
        sys.exit()


def get_exact_num_last_page(driver: webdriver, first_listing_url: str) -> int:
    """Get the number of the last page from the first listing URL."""

    print('Looking for the number of the last page:')
    driver = open_listing_url(first_listing_url, driver)
    exact_num_last_page = int(recognize_sln_selectors(driver).get_number_last_page())
    return exact_num_last_page


def open_listing_url(listing_url: str, driver: webdriver = None) -> webdriver:
    """Attempt to open a listing URL until it gets loaded."""

    while True:
        try:
            if not driver:
                driver = open_firefox_exts_headless()
            print(f'Opening URL: {listing_url}')
            driver.get(listing_url)
            break
        except TimeoutException:
            close_firefox(driver, time_exc=True)
            driver = None
            time.sleep(3)
    return driver


def parse_listing_url(driver: webdriver, listing_url: str, first_listing_url: str,
                      output_xlsx: str, add_pages: bool) -> None:
    """Get a list of all individual direct URLs of estates found in a listing URL. Terminate if no URLs found."""

    estate_urls = recognize_sln_selectors(driver).estate_urls()
    if len(estate_urls) > 0:
        print(f'Found {len(estate_urls)} estate offers on the page.\n'
              f'Crawling...\n')
        for estate_url in estate_urls:
            parse_estate_url(first_listing_url, listing_url, estate_url, driver, output_xlsx)
    else:
        if add_pages:
            raise NoAdsFound
        else:
            close_firefox(driver, no_ads_found=True)


def restart_firefox(driver: webdriver, restart: bool, listing_url: str, first_listing_url: str,
                    output_xlsx: str, add_pages: bool) -> webdriver:
    """Restart Firefox in order to avoid filling up RAM caused by Selenium memory leaks."""

    # TODO: this needs testing and probably confirming whether the allocated memory is being freed for real
    close_firefox(driver, restart=restart)
    driver = None  # destroy
    driver = open_listing_url(listing_url)
    parse_listing_url(driver, listing_url, first_listing_url, output_xlsx, add_pages)
    return driver


def close_firefox(driver: webdriver, restart: bool = False, lastp: bool = False,
                  no_ads_found: bool = False, time_exc: bool = False) -> None:
    """Close, quit, destroy the webdriver instance of Firefox and safely close the DB connection."""

    if restart:
        print('Restart of Firefox requested.\n'
              'Closing Firefox...\n')
        driver.close()
        driver.quit()
    elif time_exc:
        print('Failed to load the listing URL in time.\n'
              'Closing Firefox and retrying...\n')
        driver.close()
        driver.quit()
    elif lastp:
        print('This was the last page that was set or found in the first listing URL.\n'
              'Closing Firefox...\n')
        driver.close()
        driver.quit()
        disconnect_db()
    elif no_ads_found:
        print('No ads found on the page. Breaking out of the loop.\n'
              'Closing Firefox...\n')
        driver.close()
        driver.quit()
        disconnect_db()


# TODO: Re-design for naming the xlsx file automatically and setting output_xlsx only to False/True.
def start_crawl(listing_url: str, to_page: int = None, from_page: int = None, output_xlsx: str = None,
                lastp: bool = False, driver: webdriver = None, add_pages: bool = False, restart: bool = False,
                first_listing_url: str = None) -> None:
    """Scrape contact details from estate websites.

    :param listing_url: URL of the page with rental offers/estates.
    :param output_xlsx: Name of an output file (example.xlsx) to save the scraped data into.
    :param from_page: Begin at this page number.
    :param to_page: End at this page number.
    :param driver: Instance of Firefox Selenium webdriver.
    :param lastp: States whether the current URL is the last or not.
    :param add_pages: If True, the URLs to scrape data from were already created.
    :param restart: States whether the current instance of webdriver is going to be re-created.
    :param first_listing_url: Copy of the first listing URL (page 1).
    """

    evaluate_paging(from_page, to_page)
    if not first_listing_url:
        # to store in the database
        first_listing_url = listing_url

    if restart:
        driver = restart_firefox(driver, restart, listing_url, first_listing_url, output_xlsx, add_pages)

    if not add_pages:
        driver, from_page, lastp, to_page = \
            process_first_listing_url(add_pages, first_listing_url, from_page, lastp, listing_url, output_xlsx, to_page)
    else:
        driver = open_listing_url(listing_url, driver)
        parse_listing_url(driver, listing_url, first_listing_url, output_xlsx, add_pages)

    if lastp:
        close_firefox(driver, lastp=lastp)

    if to_page:
        process_additional_listing_urls(driver, first_listing_url, from_page, output_xlsx, to_page)


def process_first_listing_url(add_pages, first_listing_url, from_page, lastp, listing_url, output_xlsx,
                              to_page):
    driver = open_firefox_exts_headless()
    exact_num_last_page = get_exact_num_last_page(driver, first_listing_url)
    if to_page:
        if exact_num_last_page < to_page:
            print(f'The required page number {to_page} does not exist.\n'
                  f'The upper limit was lowered to the last page number {exact_num_last_page}.\n')
            to_page = exact_num_last_page
            if from_page > to_page:
                from_page = to_page
        if from_page == to_page or to_page == 1:
            print(f'Scraping only this single page number {to_page} ({exact_num_last_page} pages in total).\n')
        else:
            print(f'Scraping from a page number {from_page if from_page else "1"} '
                  f'to a page number {to_page} ({exact_num_last_page} pages in total).\n')
    else:
        print(f'Scraping from a page number {from_page if from_page else "1"} '
              f'to the last page number {exact_num_last_page}.\n')
        to_page = exact_num_last_page
    if not from_page or from_page == 1:
        print('Page number: 1 | Position: 1')
        parse_listing_url(driver, listing_url, first_listing_url, output_xlsx, add_pages)
        if from_page == to_page:
            to_page = None
            lastp = True
    return driver, from_page, lastp, to_page


def process_additional_listing_urls(driver, first_listing_url, from_page, output_xlsx, to_page):
    listing_page = recognize_sln_selectors(first_listing_url=first_listing_url, from_page=from_page,
                                           to_page=to_page).create_add_listing_urls()
    counter = 2 if not from_page else 1
    for page_num, listing_url in listing_page:
        if page_num == to_page:
            print(f"Page number: {page_num} (the last page) | Position: {counter}")
            start_crawl(listing_url, output_xlsx, driver=driver, lastp=True, add_pages=True)
        else:
            print(f"Page number: {page_num} | Position: {counter}")
            try:
                if counter % 10 == 0:
                    # this driver variable replaces the previous one that was closed
                    start_crawl(listing_url, output_xlsx, driver=driver, add_pages=True, restart=True,
                                first_listing_url=first_listing_url)
                    counter += 1
                else:
                    start_crawl(listing_url, output_xlsx, driver=driver, add_pages=True,
                                first_listing_url=first_listing_url)
                    counter += 1
            except NoAdsFound:
                close_firefox(driver, no_ads_found=True)
                break


def parse_estate_url(first_listing_url: str, listing_url: str, estate_url: str,
                     driver: webdriver, output_xlsx: str) -> None:
    """Collect required fields from the source code of an estate offer.

    :param first_listing_url: Copy of the first listing URL (page 1).
    :param listing_url: URL of the page with rental offers/estates.
    :param estate_url: URL of a single estate offer to parse.
    :param driver: Instance of Firefox.
    :param output_xlsx: Name of an output file (example.xlsx) to save the scraped data into.
    """
    refresh_count = 0
    while refresh_count < 3:
        try:
            driver.get(estate_url)
            break
        except TimeoutException:
            print('This takes too long.\n'
                  'Refreshing the page...\n')
            refresh_count += 1
    if not refresh_count == 2:
        estate = recognize_sln_selectors(driver)

        title = estate.title()
        contact_name = estate.contact_name()
        list_emails = estate.list_emails()
        list_phone_numbers = estate.list_phone_numbers()

        if list_phone_numbers or list_emails:
            if output_xlsx:
                store_in_xlsx_file(title, contact_name, list_phone_numbers, list_emails,
                                   estate_url, listing_url, first_listing_url, output_xlsx)
            else:
                store_in_database(title, contact_name, list_phone_numbers, list_emails,
                                  estate_url, listing_url, first_listing_url)
        else:
            print(f'{"Contact name not found." if contact_name == "Anonymous" else contact_name} '
                  f'| Neither a phone number or email address found.'
                  f'\n{estate_url}'
                  f'\nDiscarded\n')
            time.sleep(1)


def print_data(title: str, contact_name: str, list_phone_numbers: list, list_emails: list, estate_url: str) -> None:
    """Display a summary of collected data from an estate offer."""

    check_phones = "no phone numbers found" if not list_phone_numbers \
        else list_phone_numbers[0] + " (" + str(len(list_phone_numbers)) + " phone number(s))"
    check_emails = "no emails found" if not list_emails \
        else list_emails[0] + " (" + str(len(list_emails)) + " email(s))"
    print(f'{title} / {contact_name} / {check_phones} / {check_emails}'
          f'\n{estate_url}')


def store_in_database(title: str, contact_name: str, list_phone_numbers: list, list_emails: list,
                      estate_url: str, listing_url: str, first_listing_url: str) -> None:
    """Check whether the fields are present in the database and if not so, store them in.
    
    :param title: Title or heading of the estate offer.
    :param contact_name: First name, last name, company name, user name or any combination.
    :param list_phone_numbers: List of collected phone number from the estate offer.
    :param list_emails: List of collected email addresses from the estate offer.
    :param estate_url: URL of the estate offer. It is referred to as "direct URL" in the database.
    :param listing_url: URL of the page with rental offers/estates.
    :param first_listing_url: Copy of the first listing URL (page 1).
    """

    if not db_connected:
        connect_db()
    print_data(title, contact_name, list_phone_numbers, list_emails, estate_url)

    # the order of working with tables is contact_details, phones, emails, listing_urls, estates, first_listing_urls
    cur.execute(f"SELECT * FROM estates WHERE direct_url = '{estate_url}'")
    if cur.rowcount == 0:
        recorded_by = 'selenium_ff_scraper'
        recorded_on = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # TODO: Improve this block of assessing data for storing into DB
        if "'" in title:
            title = title.replace("'", "''")
        if "'" in contact_name:
            contact_name = contact_name.replace("'", "''")
        if len(contact_name) > 100:
            contact_name = contact_name[:100]
        if len(title) > 600:
            title = title[:600]

        # store this data in database only if the direct URL is not included yet
        cur.execute(f"SELECT * FROM contact_details WHERE contact_name = '{contact_name}'")
        if cur.rowcount == 0:
            # it does not exist, add it
            cur.execute(f"INSERT INTO contact_details (contact_name) VALUES ('{contact_name}')")
            cur.connection.commit()
            contact_id = cur.lastrowid
        else:
            #  it exists, get its id
            contact_id = cur.fetchone()[0]

        if list_phone_numbers:
            for phone in list_phone_numbers:
                cur.execute(f"SELECT * FROM phones WHERE phone = '{phone}'")
                if cur.rowcount == 0:
                    # add only if it doesn't exist
                    cur.execute(f"INSERT INTO phones (contact_id_fk, phone) VALUES ('{contact_id}', '{phone}')")
                    cur.connection.commit()

        # emails table
        if list_emails:
            for email in list_emails:
                cur.execute(f"SELECT * FROM emails WHERE email = '{email}'")
                if cur.rowcount == 0:
                    cur.execute(f"INSERT INTO emails (contact_id_fk, email) VALUES ('{contact_id}', '{email}')")
                    cur.connection.commit()

        cur.execute(f"SELECT * FROM listing_urls WHERE listing_url = '{listing_url}'")
        if cur.rowcount == 0:
            cur.execute(f"INSERT INTO listing_urls (listing_url) VALUES ('{listing_url}')")
            cur.connection.commit()
            listing_url_id = cur.lastrowid
        else:
            listing_url_id = cur.fetchone()[0]

        # I know that estate_url does not exist in DB yet, if it does, I don't want to store it anyway...
        cur.execute(f"INSERT INTO estates (contact_id_fk, listing_url_id_fk, recorded_on, recorded_by, "
                    f"title_estate, direct_url) "
                    f"VALUES ('{contact_id}', '{listing_url_id}', '{recorded_on}', '{recorded_by}', "
                    f"'{title}', '{estate_url}')")
        cur.connection.commit()
        estate_id = cur.lastrowid

        cur.execute(f"SELECT * FROM first_listing_urls WHERE first_listing_url = '{first_listing_url}'")
        if cur.rowcount == 0:
            cur.execute(f"INSERT INTO first_listing_urls (estate_id_fk, first_listing_url) "
                        f"VALUES ('{estate_id}', '{first_listing_url}')")
            cur.connection.commit()
        print('--- Storing in database ---\n')
    else:
        print('--- Already stored ---\n')


def store_in_xlsx_file(title: str, contact_name: str, list_phone_numbers: list, list_emails: list,
                       estate_url: str, listing_url: str, first_listing_url: str, output_xlsx: str) -> None:
    """Store the collected data in the XLSX file. Duplicates possible."""

    print_data(title, contact_name, list_phone_numbers, list_emails, estate_url)
    print('--- Storing in XLSX file ---\n')
    # get empty string if nothing to assign to a variable using generator comprehensions:
    phone1, phone2, phone3 = (list_phone_numbers[x] if x <= len(list_phone_numbers) - 1 else "" for x in range(3))
    email1, email2 = (list_emails[x] if x <= len(list_emails) - 1 else "" for x in range(2))
    row = [first_listing_url, listing_url, estate_url, title, contact_name, phone1, phone2, phone3, email1, email2]
    try:
        wb = load_workbook(filename=output_xlsx)
    except FileNotFoundError:
        wb = Workbook()
    ws = wb.active
    headers = ['First listing URL', 'Listing URL', 'Estate URL', 'Title', 'Contact name',
               'Phone 1', 'Phone 2', 'Phone 3', 'Email 1', 'Email 2']
    # add headers first, unless they already exist
    if ws['A1'].value != 'First listing URL':
        for index, col in enumerate(headers):
            ws.cell(1, index + 1, col)
    ws.append(row)
    wb.save(filename=output_xlsx)


begin = datetime.datetime.now()
print(f"Started at {begin.strftime('%c')}")

# TODO: add this via argprs module
# input_start_url = input('Enter an URL with a listing of ads:\n')
# output_file = input('Enter a name of an output .xlsx file (extension name included):\n')
# to_page = input('Enter a number of pages to scrape the data up to:\n')


first_url = 'https://www.traum-ferienwohnungen.de/europa/deutschland/schleswig-holstein/ergebnisse/' \
            '?person=34&is_in_clicked_search=1'
# first_url = 'https://www.milanuncios.com/alquiler-vacaciones-en-las_palmas/'
# first_url = 'https://www.vivaweek.com/fr/locations-vacances/herault-languedoc-roussillon-france/' \
#             'hebergement-type:appartement,studio,autre-appartement,bateau,catamaran,peniche,voilier,yacht,' \
#             'autre-bateau,bungalow-mobilhome,chalet,chateau-manoir,gite,insolite,cabane-arbre,moulin,phare,' \
#             'roulotte,tipi,yourte,autre-insolite,maison-villa,mas,riad,villa,autre-maison'
# first_url = 'https://vacances.seloger.com/location-vacances-france/savoie-10000007425'

xlsx = 'traum_ferienwohnungen.xlsx'

if __name__ == '__main__':
    try:
        # arguments: first_listing, to_page, from_page, output_xlsx=file.xlsx
        p = mp.Process(target=start_crawl(first_url))
        # run 'worker' in a subprocess
        p.start()
        # make the main process wait for `worker` to end
        p.join()
        # all memory used by the subprocess will be freed to the OS
    finally:
        if conn and cur:
            disconnect_db()

end = datetime.datetime.now()
print(f"Finished at {end.strftime('%c')}")
delta = str(end - begin).split('.')[0]
print(f'Time duration: {delta}')
