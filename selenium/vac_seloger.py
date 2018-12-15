import datetime
import time
import random

import pyautogui
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.common.exceptions import NoSuchElementException


class NoAdsFound(Exception):
    pass


def start_crawl(start_url, output_filename, pages=None, lastp=None, driver=None):
    """Launch Firefox and get URLs of individual ads to parse data from. Close Firefox after the last page was parsed.

    :param start_url: URL of the first page with results.
    :param output_filename: Name of the output file (example.xlsx) to save the scraped data into.
    :param pages: Number of pages of results to scrape.
    :param lastp: States whether the current URL is the last or not.
    :param driver: Instance of Firefox.
    """

    profile = FirefoxProfile(profile_directory='/home/standa/.mozilla/firefox/6v6v9nt5.vac_seloger/')
    if not driver:
        driver = webdriver.Firefox(firefox_profile=profile)
        time.sleep(2)
        # re-enable ublock and noscript "by hand"
        # TODO: find a better way around to be able to run the script in headless mode
        pyautogui.PAUSE = 0.5
        pyautogui.hotkey('winleft', 'left')
        pyautogui.hotkey('ctrl', 'shift', 'a')
        pyautogui.click(x=474, y=207, clicks=2, interval=0.25, button='left')
        pyautogui.click(x=480, y=267, clicks=2, interval=0.25, button='left')
        # this pause ensures that everything is set up properly
        time.sleep(4)

    print(f'Opening URL: {start_url}')
    driver.get(start_url)

    elements_links = driver.find_elements_by_xpath('//a[@class="vignette-link"]')
    list_of_links = [link.get_attribute('href') for link in elements_links]

    if len(list_of_links) > 0:
        print(f'Found {len(list_of_links)} ads on the page.\n'
              f'Crawling...\n')
    else:
        print('No ads found on the page.\n')
        # don't continue, terminate the script
        raise NoAdsFound

    for link in list_of_links:
        try:
            parse_add(link, driver, output_filename)
        except NoAdsFound:
            print('Closing Firefox...\n')
            driver.quit()

    if lastp:
        print('This was the last page.\n'
              'Closing Firefox...')
        driver.quit()

    # create URLs for additional pages
    if pages:
        if pages == 1:
            print('This was the last page\n'
                  'Closing Firefox...\n')
            driver.quit()
        else:
            listing_id = start_url.split('-')[-1]
            # the URL path after the the page n. 1 is different
            returned_url = f'https://vacances.seloger.com/annonces?lo={listing_id}'

            url_pages = []
            # default start of the range is at 2
            for page in range(2, pages + 1):
                url_pages.append(returned_url + f'&page={page}')
            for url_page in url_pages:
                if url_page == url_pages[-1]:
                    start_crawl(url_page, output_filename, driver=driver, lastp=True)
                else:
                    print(f"Page number {url_page.split('=')[-1]}.")
                    try:
                        start_crawl(url_page, output_filename, driver=driver)
                    except NoAdsFound:
                        print('Breaking the loop\n'
                              'Closing Firefox...\n')
                        driver.quit()
                        break


def process_data(title, seller, phone, email, ad_url, output_filename):
    scraped_list = [title, seller, phone, email, ad_url]

    print(f'{title} / {"not-found" if seller == "" else seller} / {"not-found" if phone == "" else phone} / '
          f'{"not-found" if email == "" else email}\n{ad_url}')

    if phone != '' or email != '':
        print('--- Saved ---\n')
        append_list_excel(scraped_list, output_filename)
    else:
        print('Discarded\n')


def parse_add(ad_url, driver, output_filename):
    """Get required fields from the source code of the URL and send them to save in the output .xlsx file.

    :param ad_url: URL of a single ad to parse.
    :param driver: Instance of Firefox.
    :param output_filename: Name of the output file (example.xlsx) to save the scraped data into.
    """

    driver.get(ad_url)
    # the list has to contain the field in this order: titulo, contacto, telefono, email, url

    try:
        # select the one which is True (not an empty string) using OR operator
        seller = driver.find_element_by_xpath('//p[@class="title"]').text or \
                 driver.find_elements_by_xpath('//p[@class="title"]')[1].text
    except NoSuchElementException:
        seller = 'not-found'

    if seller == 'Amivac.com':
        # TODO: this was not tested yet!!!
        parse_amivac_add(driver, output_filename)
    elif seller == 'Villes':
        print('The IP address was blocked :(')
        raise NoAdsFound
    else:
        try:
            title = driver.find_element_by_xpath('//h1[@itemprop="name"]').text
        except NoSuchElementException:
            title = 'not-found'

        # TODO: try to get any email address from I don't know where...
        email = ''

        try:
            popup = driver.find_element_by_xpath('//a[@class="jsInfosProfilTelPopin"]').click()
        except NoSuchElementException:
            popup = 'not-found'

        if popup != 'not-found':
            time.sleep(1)

            try:
                phone1 = driver.find_element_by_xpath('//div[@class="tel-big"]').text
            except NoSuchElementException:
                phone1 = ''

            try:
                phone2 = driver.find_elements_by_xpath('//div[@class="tel-big"]')[1].text
            except IndexError:
                phone2 = ''

            if phone2 != '':
                phone = phone1 + ', ' + phone2
            else:
                phone = phone1

            process_data(title, seller, phone, email, ad_url, output_filename)
        else:
            print(f'{seller} | The phone number detail not found.\n{ad_url}\nDiscarded\n')
            time.sleep(1)


def random_wait():
    wait = random.randint(3, 6)
    print(f'Waiting {wait} seconds before scraping another add.\n')
    time.sleep(wait)


def parse_amivac_add(driver, output_filename):
    # open the amivac add
    # this was tested only separately, not as a part of the loop, getting to the Amivac page from vacances.seloger.com
    # what includes transferring phase
    driver.find_elements_by_xpath('//button[@id="en-savoir-plus"]')[1].click()
    time.sleep(3)

    try:
        title = driver.find_element_by_xpath('//strong[@class="entete6"]').text
    except NoSuchElementException:
        title = 'not-found'

    try:
        seller = driver.find_element_by_xpath('//tr[16]').text.splitlines()[0]
    except NoSuchElementException:
        seller = 'not-found'

    # show phone numbers
    driver.find_element_by_xpath('//a[@class="showtel"]').click()

    try:
        phone = ', '.join([number.split(':')[-1].strip() for number in
                           driver.find_element_by_xpath('//span[@id="tels"]').text.splitlines() if
                           'Télécopie' not in number])
    except NoSuchElementException:
        phone = 'not-found'

    email = ''

    ad_url = driver.current_url

    process_data(title, seller, phone, email, ad_url, output_filename)


def append_list_excel(scraped_list, output_filename):
    """Open the excel spreadsheet if it exists, if not, create it.
    Write headers if not present and append scraped data from an individual ad to the excel spreadsheet.

    :param scraped_list: Scraped data as a list.
    :param output_filename: Name of the output file (example.xlsx) to save the scraped data into.
    """

    try:
        wb = load_workbook(filename=output_filename)
    except FileNotFoundError:
        wb = Workbook()
    ws = wb.active
    headers = ['titulo', 'contacto', 'telefono', 'email', 'url']
    if ws['A1'].value != 'titulo':
        for index, col in enumerate(headers):
            ws.cell(1, index + 1, col)
    ws.append(scraped_list)
    wb.save(filename=output_filename)


# start_url = input('Enter an URL with filtered ads:\n')

begin = datetime.datetime.now()
print(f"Started at {begin.strftime('%c')}")

input_start_url = 'https://vacances.seloger.com/location-vacances-france/savoie-10000007425'
output_file = 'vac_seloger.xlsx'

start_crawl(input_start_url, output_file, 400)

end = datetime.datetime.now()
print(f"Finished at {end.strftime('%c')}")
delta = str(end - begin).split('.')[0]
print(f'Time duration: {delta}')
