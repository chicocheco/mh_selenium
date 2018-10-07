import json
import os
import sys

from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.firefox.firefox_profile import AddonFormatError
from selenium.common.exceptions import NoSuchElementException, InvalidSessionIdException


# temporal fix for webdriver.FirefoxProfile()
class FirefoxProfileWithWebExtensionSupport(webdriver.FirefoxProfile):
    def _addon_details(self, addon_path):
        try:
            return super()._addon_details(addon_path)
        except AddonFormatError:
            try:
                with open(os.path.join(addon_path, 'manifest.json'), 'r') as f:
                    manifest = json.load(f)
                    return {
                        'id': manifest['applications']['gecko']['id'],
                        'version': manifest['version'],
                        'name': manifest['name'],
                        'unpack': False,
                    }
            except (IOError, KeyError) as e:
                raise AddonFormatError(str(e), sys.exc_info()[2])


def start_crawl(start_url, output_filename, pages=None, lastp=None, driver=None):
    """Launch Firefox and get URLs of individual ads to parse data from. Close Firefox after the last page was parsed.

    :param start_url: URL of the first page with results.
    :param output_filename: Name of the output file (example.xlsx) to save the scraped data into.
    :param pages: Number of pages of results to scrape.
    :param lastp: Declare whether the current URL is the last or not.
    :param driver: Instance of Firefox.
    """

    profile = FirefoxProfileWithWebExtensionSupport()
    profile.set_preference('permissions.default.image', 2)
    profile.add_extension(extension='/home/standa/PycharmProjects/mh/selenium/noscript.xpi')
    profile.update_preferences()

    if not driver:
        driver = webdriver.Firefox(firefox_profile=profile)
    print(f'Opening URL: {start_url}')
    driver.get(start_url)
    driver.minimize_window()

    elements_links = driver.find_elements_by_xpath('//a[@class="re-Card-link"]')
    list_of_links = [link.get_attribute('href') for link in elements_links]
    if len(list_of_links) > 0:
        print(f'Found {len(list_of_links)} ads on the page.\n'
              f'Crawling...\n')
    else:
        print('No ads found on the page.\n')

    for link in list_of_links:
        parse_add(link, driver, output_filename)

    if lastp:
        print('This was the last page\n'
              'Closing Firefox...')
        driver.close()

    # create URLs for additional pages
    if pages:
        url_pages = []
        for page in range(2, pages + 1):
            url_pages.append(input_start_url.replace('l?', f'l/{page}?'))
        for url_page in url_pages:
            if url_page == url_pages[-1]:
                start_crawl(url_page, output_filename, driver=driver, lastp=True)
            else:
                start_crawl(url_page, output_filename, driver=driver)


def parse_add(ad_url, driver, output_filename):
    """Get required fields from the source code of the URL and send them to save in the output .xlsx file.

    :param ad_url: URL of a single ad to parse.
    :param driver: Instance of Firefox.
    :param output_filename: Name of the output file (example.xlsx) to save the scraped data into.
    """

    driver.get(ad_url)
    scraped_list = []

    # TODO: safely skip dead links
    try:
        title = driver.find_element_by_xpath('//h1[@class="property-title"]').text
        if title == '#':
            title = 'not-found'
    except NoSuchElementException:
        title = 'not-found'
    scraped_list.append(title)

    try:
        seller = driver.find_element_by_xpath('//span[@id="ctl00_ucInfoRequestGeneric_divContact"]').text[10:]
    except NoSuchElementException:
        seller = 'not-found'
    scraped_list.append(seller)

    try:
        phone = driver.find_element_by_name('ctl00$hid_AdPhone').get_attribute('value')
        if phone == '#':
            phone = 'not-found'
        else:
            phone = f'{phone[4:7]} {phone[7:10]} {phone[10:13]}'
    except NoSuchElementException:
        phone = 'not-found'
        # TODO: phone regex from description
    scraped_list.append(phone)

    # TODO: email regex from description
    email = 'not-found'
    scraped_list.append(email)

    scraped_list.append(ad_url)



    print(f'{title} / {seller} / {phone} / {email}\n{ad_url}')

    if phone != 'not-found' or email != 'not-found':
        print('--- Saved ---\n')
        append_list_excel(scraped_list, output_filename)
    else:
        print('Discarded\n')


def append_list_excel(scraped_list, output_filename):
    """Open the excel spreadsheet if it exists, if not, create it.
    Write headers if not present and append scraped data from an individual ad to the excel spreadsheet.

    :param scraped_list: Scraped data as a list.
    :param output_filename: Name of the output file (example.xlsx) to save the scraped data into.
    """

    try:
        wb = load_workbook(filename=output_filename)
    except FileNotFoundError:
        wb = Workbook()
    ws = wb.active
    headers = ['titulo', 'contacto', 'telefono', 'email', 'url']
    if ws['A1'].value != 'titulo':
        for index, col in enumerate(headers):
            ws.cell(1, index + 1, col)
    ws.append(scraped_list)
    wb.save(filename=output_filename)


# start_url = input('Enter an URL with filtered ads:\n')


input_start_url = 'https://www.fotocasa.es/es/alquiler/casas/ceuta-provincia/todas-las-zonas/l' \
                  '?latitude=40&longitude=-4&combinedLocationIds=724,10,51,0,0,0,0,0,0&gridType=3'
output_file = 'fotocasa_test.xlsx'

start_crawl(input_start_url, output_file, 3)

# TODO: delete duplicates from the output
