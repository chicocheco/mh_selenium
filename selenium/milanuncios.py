import json
import os
import sys
import datetime
import time
import pyautogui

from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.firefox.firefox_profile import AddonFormatError
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.common.exceptions import NoSuchElementException


# temporal fix for webdriver.FirefoxProfile()
class FirefoxProfileWithWebExtensionSupport(webdriver.FirefoxProfile):
    def _addon_details(self, addon_path):
        try:
            return super()._addon_details(addon_path)
        except AddonFormatError:
            try:
                with open(os.path.join(addon_path, 'manifest.json'), 'r') as f:
                    manifest = json.load(f)
                    return {
                        'id': manifest['applications']['gecko']['id'],
                        'version': manifest['version'],
                        'name': manifest['name'],
                        'unpack': False,
                    }
            except (IOError, KeyError) as e:
                raise AddonFormatError(str(e), sys.exc_info()[2])


def start_crawl(start_url, output_filename, pages=None, lastp=None, driver=None):
    """Launch Firefox and get URLs of individual ads to parse data from. Close Firefox after the last page was parsed.

    :param start_url: URL of the first page with results.
    :param output_filename: Name of the output file (example.xlsx) to save the scraped data into.
    :param pages: Number of pages of results to scrape.
    :param lastp: States whether the current URL is the last or not.
    :param driver: Instance of Firefox.
    """

    profile = FirefoxProfile(profile_directory='/home/standa/.mozilla/firefox/zumag6xp.wd_milanuncios/')
    if not driver:
        driver = webdriver.Firefox(firefox_profile=profile, options=options)
        time.sleep(2)
        # re-enable ublock and noscript "by hand"
        pyautogui.PAUSE = 0.5
        pyautogui.hotkey('winleft', 'left')
        pyautogui.hotkey('ctrl', 'shift', 'a')
        pyautogui.click(x=474, y=207, clicks=2, interval=0.25, button='left')
        pyautogui.click(x=480, y=267, clicks=2, interval=0.25, button='left')

    print(f'Opening URL: {start_url}')
    driver.get(start_url)

    elements_links = driver.find_elements_by_xpath('//a[@class="aditem-detail-title"]')
    list_of_links = [link.get_attribute('href') for link in elements_links]

    if len(list_of_links) > 0:
        print(f'Found {len(list_of_links)} ads on the page.\n'
              f'Crawling...\n')
    else:
        print('No ads found on the page.\n')
        # TODO: How to break out of the loop??

    for link in list_of_links:
        parse_add(link, driver, output_filename)

    if lastp:
        print('This was the last page.\n'
              'Closing Firefox...')
        driver.quit()

    # create URLs for additional pages
    if pages:
        if pages == 1:
            print('This was the last page\n'
                  'Closing Firefox...')
            driver.quit()
        else:
            url_pages = []
            for page in range(18, pages + 1):
                url_pages.append(input_start_url + f'?pagina={page}')
            for url_page in url_pages:
                if url_page == url_pages[-1]:
                    start_crawl(url_page, output_filename, driver=driver, lastp=True)
                else:
                    print(f"Page number {url_page.split('=')[-1]}.")
                    start_crawl(url_page, output_filename, driver=driver)


def parse_add(ad_url, driver, output_filename):
    """Get required fields from the source code of the URL and send them to save in the output .xlsx file.

    :param ad_url: URL of a single ad to parse.
    :param driver: Instance of Firefox.
    :param output_filename: Name of the output file (example.xlsx) to save the scraped data into.
    """

    driver.get(ad_url)
    scraped_list = []

    try:
        title = driver.find_element_by_xpath('//div[@class="pagAnuTituloBox"]/a').text
    except NoSuchElementException:
        title = 'not-found'

    scraped_list.append(title)
    scraped_list.append(ad_url)

    ad_id = ad_url.split('/')[-1][-13:-4]
    seller_url = f'https://www.milanuncios.com/datos-contacto/?usePhoneProxy=0&from=detail&id={ad_id}'
    driver.get(seller_url)

    try:
        seller = driver.find_element_by_xpath('//strong').text
    except NoSuchElementException:
        seller = 'not-found'
    scraped_list.insert(1, seller)

    try:
        phone1 = driver.find_element_by_xpath('//div[@class="telefonos"]').text
    except NoSuchElementException:
        phone1 = 'not-found'

    try:
        phone2 = driver.find_elements_by_xpath('//div[@class="telefonos"]')[1].text
    except IndexError:
        phone2 = 'not-found'

    if phone2 != 'not-found':
        phone = phone1 + ', ' + phone2
    else:
        phone = phone1

    scraped_list.insert(2, phone)

    # TODO: email regex from description
    email = 'not-found'
    scraped_list.insert(3, email)

    print(f'{title} / {seller} / {phone} / {email}\n{ad_url}')

    if phone != 'not-found' or email != 'not-found':
        print('--- Saved ---\n')
        append_list_excel(scraped_list, output_filename)
    else:
        print('Discarded\n')


def append_list_excel(scraped_list, output_filename):
    """Open the excel spreadsheet if it exists, if not, create it.
    Write headers if not present and append scraped data from an individual ad to the excel spreadsheet.

    :param scraped_list: Scraped data as a list.
    :param output_filename: Name of the output file (example.xlsx) to save the scraped data into.
    """

    try:
        wb = load_workbook(filename=output_filename)
    except FileNotFoundError:
        wb = Workbook()
    ws = wb.active
    headers = ['titulo', 'contacto', 'telefono', 'email', 'url']
    if ws['A1'].value != 'titulo':
        for index, col in enumerate(headers):
            ws.cell(1, index + 1, col)
    ws.append(scraped_list)
    wb.save(filename=output_filename)


# start_url = input('Enter an URL with filtered ads:\n')

input_start_url = 'https://www.milanuncios.com/alquiler-vacaciones-en-las_palmas/'
output_file = 'milanuncios.xlsx'

begin = datetime.datetime.now()
print(f"Started at {begin.strftime('%c')}")

start_crawl(input_start_url, output_file, 30)

end = datetime.datetime.now()
print(f"Finished at {end.strftime('%c')}")
delta = str(end - begin).split('.')[0]
print(f'Time duration: {delta}')


# TODO: delete duplicates from the output
