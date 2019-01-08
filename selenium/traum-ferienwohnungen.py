import multiprocessing as mp
import datetime
import time
import json
import os
import sys

import pyautogui
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.webdriver.firefox.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.firefox.firefox_profile import AddonFormatError
from selenium.common.exceptions import TimeoutException, InvalidElementStateException, ElementNotInteractableException
from urllib3.exceptions import NewConnectionError, MaxRetryError


class NoAdsFound(Exception):
    pass


# in case of installing modern extensions:
class FirefoxProfileWithWebExtensionSupport(webdriver.FirefoxProfile):
    def _addon_details(self, addon_path):
        try:
            return super()._addon_details(addon_path)
        except AddonFormatError:
            try:
                with open(os.path.join(addon_path, 'manifest.json'), 'r') as f:
                    manifest = json.load(f)
                    return {
                        'id': manifest['applications']['gecko']['id'],
                        'version': manifest['version'],
                        'name': manifest['name'],
                        'unpack': False,
                    }
            except (IOError, KeyError) as e:
                raise AddonFormatError(str(e), sys.exc_info()[2])


def open_firefox_exts_manually():
    """Launch Firefox and enable add-ons on the screen."""

    profile = FirefoxProfileWithWebExtensionSupport(
        profile_directory='/home/standa/.mozilla/firefox/mrgxem71.scraping/')
    driver = webdriver.Firefox(firefox_profile=profile)
    # driver.maximize_window()
    time.sleep(1)
    pyautogui.PAUSE = 1
    pyautogui.hotkey('winleft', 'left')
    pyautogui.hotkey('ctrl', 'shift', 'a')
    # disable and enable ublock and noscript
    # pyautogui.center(pyautogui.locateOnScreen('./selenium/enable_button.png'))
    pyautogui.click(x=604, y=270, clicks=1, interval=0.25, button='left')
    pyautogui.click(x=604, y=360, clicks=1, interval=0.25, button='left')
    driver.set_page_load_timeout(15)
    time.sleep(1)
    return driver


def open_firefox_exts_headless():
    """Launch Firefox in headless mode, enable add-ons."""

    print('Launching Firefox in headless mode...')
    profile = FirefoxProfile(profile_directory='/home/standa/.mozilla/firefox/mrgxem71.scraping/')
    options = Options()
    options.headless = True
    driver = webdriver.Firefox(firefox_profile=profile, options=options)
    time.sleep(2)
    while True:
        try:
            driver.get('about:addons')
            driver.find_element_by_name('NoScript').click()
            driver.find_element_by_id('detail-enable-btn').click()
            print(' - NoScript add-on enabled.')
            driver.back()
            driver.find_element_by_name('uBlock Origin').click()
            driver.find_element_by_id('detail-enable-btn').click()
            print(' - uBlock Origin add-on enabled.\n')
            break
        except (NoSuchElementException, InvalidElementStateException):
            print('Could not enable the extensions. Retrying...')
            time.sleep(2)
    driver.set_page_load_timeout(30)
    return driver


def start_crawl(start_url, output_filename, to_page=None, from_page=None, lastp=None, driver=None, add_pages=None,
                restart=None):
    """Launch Firefox and get URLs of individual ads to parse data from.

    :param start_url: URL of the first page with results.
    :param output_filename: Name of the output file (example.xlsx) to save the scraped data into.
    :param from_page: Begin at this page of the search listing. It must be higher than 1.
    :param to_page: End at this page of the search listing.
    :param driver: WebDriver instance of Firefox.
    :param lastp: States whether the current URL is the last or not.
    :param add_pages: States whether the functions is run within a loop going over additional pages.
    :param restart: States whether the current instance of WebDriver is going to be re-created.
    """

    if from_page and not to_page:
        print('Set the upper limit of the pages to scrape.\n'
              'Terminating...')
        sys.exit()
    if to_page and from_page:
        if to_page < from_page:
            print('Set the upper limit higher than the lower limit of the pages to scrape.\n'
                  'Terminating...')
            sys.exit()

    if restart:
        print('Closing Firefox to free up the memory...\n')
        driver.close()
        driver.quit()
        driver = None  # destroy the driver

    # TODO: test 'and' part
    if from_page and from_page > 1:
        driver = open_firefox_exts_headless()
    else:
        while True:
            try:
                if not driver:
                    driver = open_firefox_exts_headless()
                print(f'Opening URL: {start_url}')
                driver.get(start_url)
                break
            except TimeoutException:
                print('This takes too long.\n'
                      'Closing Firefox and retrying...\n')
                driver.close()
                driver.quit()
                driver = None  # destroy the driver
                time.sleep(3)
                driver = open_firefox_exts_headless()
            except (ConnectionRefusedError, NewConnectionError, MaxRetryError):
                print('An exception occurred.\n'
                      'Closing Firefox and retrying in 10 seconds...\n')
                if driver:
                    driver.close()
                    driver.quit()
                    driver = None  # destroy the driver
                time.sleep(10)
                driver = open_firefox_exts_headless()

        elements_links = driver.find_elements_by_xpath('//a[@class="js-link"]')
        list_of_links = [link.get_attribute('href') for link in elements_links]

        if len(list_of_links) > 0:
            print(f'Found {len(list_of_links)} ads on the page.\n'
                  f'Crawling...\n')
        else:
            print('No ads found on the page.\n')
            raise NoAdsFound

        for link in list_of_links:
            try:
                parse_add(link, driver, output_filename)
            except NoAdsFound:
                print('Closing Firefox...\n')
                driver.quit()
                driver = None
    if lastp:
        print('This was the last page (1).\n'
              'Closing Firefox...')
        driver.close()
        driver.quit()

    if restart:
        return driver

    # create URLs for additional pages, if to_page is None, it terminates with the else block
    if to_page:
        if to_page == 1:
            print('This was the last page (2)\n'
                  'Closing Firefox...\n')
            driver.quit()
            driver = None
        else:
            url_pages = {}
            # default start of the range is at 2 and that's when the URL pattern changes
            for page in range(from_page or 2, to_page + 1):
                url_chunks = start_url.split('/')
                url_chunks[-2] = f'seite-{page}'
                url_pages[page] = '/'.join(url_chunks)
            counter = 2 if not from_page else 1
            for number, url_page in url_pages.items():
                if number == max(url_pages.keys()):
                    start_crawl(url_page, output_filename, driver=driver, lastp=True, add_pages=True)
                else:
                    print(f"Page number: {number} | Position: {counter}")
                    try:
                        if counter % 10 == 0:
                            # this driver variable replaces the previous one that was closed
                            driver = start_crawl(url_page, output_filename, driver=driver, add_pages=True, restart=True)
                            counter += 1
                        else:
                            start_crawl(url_page, output_filename, driver=driver, add_pages=True)
                            counter += 1
                    except NoAdsFound:
                        print('Breaking out of the loop\n'
                              'Closing Firefox...\n')
                        driver.close()
                        driver.quit()
                        break
    if not add_pages:
        print('This was the last page (3)\n'
              'Closing Firefox...\n')
        driver.close()
        driver.quit()


def process_data(title, seller, phone, email, ad_url, output_filename):
    """Print out the collected data, save or discard them based on the presence of a phone number or email."""

    scraped_list = [title, seller, phone, email, ad_url]

    print(f'{title} / {"not-found" if seller == "" else seller} / {"not-found" if phone == "" else phone} / '
          f'{"not-found" if email == "" else email}\n{ad_url}')

    if phone != '' or email != '':
        print('--- Saved ---\n')
        append_list_excel(scraped_list, output_filename)
    else:
        print('Discarded\n')


def click_open(driver):
    """Open the items on the website when it is displayed in a narrow window/phone version:

    :param driver: Selenium webdriver.
    """

    try:
        driver.find_element_by_xpath('//a[@id="mod-mobileAccordionToggle-5"]').click()
        time.sleep(0.2)
    except (NoSuchElementException, ElementNotInteractableException):
        print('Could not click on the "Kontakt" button.')
    try:
        driver.find_element_by_xpath('//a[@id="mod-mobileAccordionToggle-8"]').click()
        time.sleep(0.2)
    except (NoSuchElementException, ElementNotInteractableException):
        print('Could not click on the "Weitere Informationen" button.')


def parse_add(ad_url, driver, output_filename):
    """Get required fields from the source code of the URL and send them to save in the output .xlsx file.

    :param ad_url: URL of a single ad to parse.
    :param driver: Instance of Firefox.
    :param output_filename: Name of the output file (example.xlsx) to save the scraped data into.
    """

    # TODO: test to catch a possible TimeoutException
    while True:
        try:
            driver.get(ad_url)
            break
        except TimeoutException:
            print('This takes too long.\n'
                  'Refreshing the page...\n')
        except (ConnectionRefusedError, NewConnectionError, MaxRetryError):
            print('An exception occurred.\n'
                  'Refreshing the page in 10 seconds...\n')
            time.sleep(10)

    # click_open(driver)

    # the list has to contain the field in this order: titulo, contacto, telefono, email, url
    try:
        title = driver.find_element_by_xpath('//div[@class="df t--fd-c m--fd-c"]'
                                             '/span[@class="pl-s t--p-0 t--light m--p-0 m--light"]').text
    except NoSuchElementException:
        title = 'not-found'
    try:
        seller = driver.find_element_by_xpath('//p[@class="fs-xl bold mb-s mt-s m--tac"]').text
    except NoSuchElementException:
        seller = 'not-found'
    # TODO: try to get any email address from I don't know where...
    email = ''

    try:
        popup = driver.find_element_by_xpath('//a[@class="us-none fs-m df ai-c"]').click()
    except NoSuchElementException:
        popup = 'not-found'

    if popup != 'not-found':
        time.sleep(1)
        try:
            phone1 = driver.find_element_by_xpath('//li[@class="mb-s"]').text
        except NoSuchElementException:
            phone1 = ''
        try:
            phone2 = driver.find_elements_by_xpath('//li[@class="mb-s"]')[1].text
        except IndexError:
            phone2 = ''
        if phone2 != '':
            phone = phone1 + ', ' + phone2
        else:
            phone = phone1
        process_data(title, seller, phone, email, ad_url, output_filename)
    else:
        print(f'{seller} | The phone number detail not found.\n{ad_url}\nDiscarded\n')
        time.sleep(1)


# TODO: develop a function to work with a database
def append_list_excel(scraped_list, output_filename):
    """Create new or open an existing excel spreadsheet.
    Write headers if not present and append scraped data from an individual ad to the excel spreadsheet.

    :param scraped_list: Scraped data as a list.
    :param output_filename: Name of the output file (example.xlsx) to save the scraped data into.
    """

    try:
        wb = load_workbook(filename=output_filename)
    except FileNotFoundError:
        wb = Workbook()
    ws = wb.active
    headers = ['titulo', 'contacto', 'telefono', 'email', 'url']

    # add headers first, unless they already exist
    if ws['A1'].value != 'titulo':
        for index, col in enumerate(headers):
            ws.cell(1, index + 1, col)
    ws.append(scraped_list)
    wb.save(filename=output_filename)


begin = datetime.datetime.now()
print(f"Started at {begin.strftime('%c')}")

# input_start_url = input('Enter an URL with a listing of ads:\n')
# output_file = input('Enter a name of an output .xlsx file (extension name included):\n')
# to_page = input('Enter a number of pages to scrape the data up to:\n')

input_start_url = 'https://www.traum-ferienwohnungen.de/europa/deutschland/schleswig-holstein/ergebnisse/' \
                  '?person=34&is_in_clicked_search=1'
output_file = 'traum_ferienwohnungen.xlsx'
pages = 200


# start_crawl(input_start_url, output_file, 700, from_page=619)
# start_crawl(input_start_url, output_file)

if __name__ == '__main__':
    p = mp.Process(target=start_crawl(input_start_url, output_file, 700, from_page=102))
    # run 'worker' in a subprocess
    p.start()
    # make the main process wait for `worker` to end
    p.join()
    # all memory used by the subprocess will be freed to the OS

end = datetime.datetime.now()
print(f"Finished at {end.strftime('%c')}")
delta = str(end - begin).split('.')[0]
print(f'Time duration: {delta}')
